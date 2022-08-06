from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd
from main import main_foo, find_news
from operator import itemgetter
from datetime import datetime
import re


def news_from_excel():
    wb = load_workbook(filename='news.xlsx')
    ws = wb['Statistic']
    list_from_excel = []
    temp = []
    for i in range(2, ws.max_row):
        for j in range(2, ws.max_column+1):
            if j != 3: temp.append(ws.cell(row=i, column=j).value)
            else:
                date = ws.cell(row=i, column=j).value
                if date == 'б/д':
                    temp.append('0.0.0')
                else:
                    date = f'{date}'.split(' ')
                    if re.search(r'Jan', date[1]):
                        month = '01'
                    elif re.search(r'Feb', date[1]):
                        month = '02'
                    elif re.search(r'Mar', date[1]):
                        month = '03'
                    elif re.search(r'Apr', date[1]):
                        month = '04'
                    elif re.search(r'May', date[1]):
                        month = '05'
                    elif re.search(r'Jun', date[1]):
                        month = '06'
                    elif re.search(r'Jul', date[1]):
                        month = '07'
                    elif re.search(r'Aug', date[1]):
                        month = '08'
                    elif re.search(r'Sep', date[1]):
                        month = '09'
                    elif re.search(r'Oct', date[1]):
                        month = '10'
                    elif re.search(r'Nov', date[1]):
                        month = '11'
                    elif re.search(r'Dec', date[1]):
                        month = '12'
                    else:
                        month = 'no data'
                    date = datetime(int(date[2]), int(month), int(date[0]))
                    date = date.strftime('%Y.%m.%d')
                    temp.append(date)
        list_from_excel.append(temp)
        temp = []

    return list_from_excel


def advert_from_excel():
    wb = load_workbook(filename='ads.xlsx')
    ws = wb['Statistic']
    list_from_excel = []
    temp = []
    for i in range(2, ws.max_row):
        for j in range(2, ws.max_column+1):
            if j != 3: temp.append(ws.cell(row=i, column=j).value)
            else:
                date = ws.cell(row=i, column=j).value
                date = f'{date}'.split(' ')
                if re.search(r'Jan', date[1]):
                    month = '01'
                elif re.search(r'Feb', date[1]):
                    month = '02'
                elif re.search(r'Mar', date[1]):
                    month = '03'
                elif re.search(r'Apr', date[1]):
                    month = '04'
                elif re.search(r'May', date[1]):
                    month = '05'
                elif re.search(r'Jun', date[1]):
                    month = '06'
                elif re.search(r'Jul', date[1]):
                    month = '07'
                elif re.search(r'Aug', date[1]):
                    month = '08'
                elif re.search(r'Sep', date[1]):
                    month = '09'
                elif re.search(r'Oct', date[1]):
                    month = '10'
                elif re.search(r'Nov', date[1]):
                    month = '11'
                elif re.search(r'Dec', date[1]):
                    month = '12'
                else:
                    month = '0.0.0'

                date = datetime(int(date[2]), int(month), int(date[0]))
                date = date.strftime('%Y.%m.%d')
                temp.append(date)
        list_from_excel.append(temp)
        temp = []

    return list_from_excel


def excel_news(period, *args):

    if period == 'm':
        list_result = find_news('m')
    else:
        list_result = find_news('p', *args)

    news_from_excel = news_from_excel()
    flag = 0
    for i in list_result:
        for j in news_from_excel:
            if i[2] == j[2]: flag += 1
        if flag == 0: news_from_excel.append(i)
        else: flag = 0

    news_from_excel = sorted(news_from_excel, key=lambda d: tuple(map(int, d[1].split('.'))))

    df = pd.DataFrame({'№': [], 'Имя запроса': [], 'Дата': [], 'URL объявления': [], 'Описание': []})
    df.to_excel('news.xlsx', sheet_name='Statistic', index=False)

    wb = load_workbook(filename='news.xlsx')
    ws = wb['Statistic']
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(4)].width = 50
    ws.column_dimensions[get_column_letter(5)].width = 50
    row = 2
    for j in news_from_excel:
        ws[f'A{row}'] = f'{row - 1}'
        ws[f'B{row}'] = f'{j[0]}'
        if j[1] != '0.0.0':
            temp = f'{j[1]}'.split('.')
            date = datetime(int(temp[0]), int(temp[1]), int(temp[2]))
            date = date.strftime('%d %B %Y')
            ws[f'C{row}'] = f'{date}'
        else:
            ws[f'C{row}'] = 'б/д'
        # date_public = date_public.strftime('%Y.%m.%d')
        ws[f'D{row}'] = f'{j[2]}'
        ws[f'E{row}'] = f'{j[3]}'
        row += 1

    # for j in list_result:
    #     count = True
    #     for i in ws['D']:
    #         if j[2] == i.value:
    #             count = False
    #             break
    #     if count:
    #         ws[f'A{row}'] = f'{row - 1}'
    #         ws[f'B{row}'] = f'{j[0]}'
    #         ws[f'C{row}'] = f'{j[1]}'
    #         ws[f'D{row}'] = f'{j[2]}'
    #         ws[f'E{row}'] = f'{j[3]}'
    #         row += 1

    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_headers = Alignment(horizontal='center', vertical='center')
    for cell in ws['A']:
        cell.alignment = alignment
    for cell in ws['B']:
        cell.alignment = alignment
    for cell in ws['C']:
        cell.alignment = alignment
    for cell in ws['D']:
        cell.alignment = alignment
    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws['A1'].alignment = alignment_headers
    ws['B1'].alignment = alignment_headers
    ws['C1'].alignment = alignment_headers
    ws['D1'].alignment = alignment_headers
    ws['E1'].alignment = alignment_headers

    for k in range(1, row):
        ws.row_dimensions[k].height = 40
    wb.save('news.xlsx')
    wb.close()
    return


def excel():

    list_result = main_foo()
    data_from_excel = advert_from_excel()

    flag = 0
    for i in list_result:
        for j in data_from_excel:
            if i[2] == j[2]: flag += 1
        if flag == 0:
            data_from_excel.append(i)
        else:
            flag = 0

    data_from_excel = sorted(data_from_excel, key=lambda d: tuple(map(int, d[1].split('.'))))

    df = pd.DataFrame({'№': [], 'Имя запроса': [], 'Дата': [], 'URL объявления': [], 'Описание': []})
    df.to_excel('ads.xlsx', sheet_name='Statistic', index=False)

    wb = load_workbook(filename='ads.xlsx')
    ws = wb['Statistic']
    row = ws.max_row + 1
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(4)].width = 50
    ws.column_dimensions[get_column_letter(5)].width = 50

    for j in data_from_excel:
        ws[f'A{row}'] = f'{row - 1}'
        ws[f'B{row}'] = f'{j[0]}'
        temp = f'{j[1]}'.split('.')
        date = datetime(int(temp[0]), int(temp[1]), int(temp[2]))
        date = date.strftime('%d %B %Y')
        ws[f'C{row}'] = f'{date}'
        ws[f'D{row}'] = j[2]
        ws[f'E{row}'] = f'{j[3]}'
        row += 1


    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_headers = Alignment(horizontal='center', vertical='center')
    for cell in ws['A']:
        cell.alignment = alignment
    for cell in ws['B']:
        cell.alignment = alignment
    for cell in ws['C']:
        cell.alignment = alignment
    for cell in ws['D']:
        cell.alignment = alignment
    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws['A1'].alignment = alignment_headers
    ws['B1'].alignment = alignment_headers
    ws['C1'].alignment = alignment_headers
    ws['D1'].alignment = alignment_headers
    ws['E1'].alignment = alignment_headers

    for k in range(1, row):
        ws.row_dimensions[k].height = 40
    wb.save('ads.xlsx')
    wb.close()
    return


if __name__ == '__main__':
    excel()
    # excel_news('p', '1/1/2021', '5/31/2021')
    # excel_news('p', '1/1/2022', '1/31/2022')
    # excel_news('m')
    # date_from_excel()